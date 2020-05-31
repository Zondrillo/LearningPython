import pandas as pd
import numpy as np
import time
import os
import xlsxwriter as xl
from openpyxl import load_workbook


def pivot(factories, budget):  # создаёт списки сводных таблиц для каждого грузополучателя, в соответствии с бюджетом
    df = []
    for factory in factories:
        temp_pivot = data_pt.query(f'Завод == ["{factory}"] & Раздел_ГКПЗ == ["{budget}"]')
        if temp_pivot.size != 0:
            df.append(temp_pivot)
    return df


def big_table(ws):
    lst = []
    i = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=19):
        tmp_lst = [i]
        j = 0
        while j < len(row):
            if j == 3:
                tmp_lst.append(None)
            tmp_lst.append(row[j].value)
            j += 1
        lst.append(tmp_lst)
        i += 1
    return lst


def add_big_table(lst, ws, row_num=8):
    r_num = row_num
    format1 = final_wb.add_format({'align': 'center', 'border': True, 'font': 'Tahoma', 'font_size': 16,
                                   'text_wrap': True})
    format1.set_align('vcenter')
    quantity_format = final_wb.add_format({'num_format': '#,##0.00', 'align': 'center', 'border': True,
                                           'font': 'Tahoma', 'font_size': 16})
    quantity_format.set_align('vcenter')
    for row in lst:
        ws.write_row(f'A{r_num}', row[:6], format1)
        ws.write_formula(r_num - 1, 6, f'=SUM(H{r_num}:T{r_num})', quantity_format)
        ws.write_row(f'H{r_num}', row[6:], quantity_format)
        r_num += 1


def make_head(ws):
    format1 = final_wb.add_format({'align': 'right', 'italic': True, 'font': 'Tahoma', 'font_size': 16})
    merge_format1 = final_wb.add_format({'align': 'center', 'border': 1, 'font': 'Tahoma', 'font_size': 16,
                                         'text_wrap': True})
    merge_format1.set_align('vcenter')
    merge_format2 = final_wb.add_format({'align': 'center', 'bold': True, 'font': 'Tahoma', 'font_size': 16})
    merge_format2.set_align('vcenter')
    merge_format3 = final_wb.add_format({'align': 'center', 'font': 'Tahoma', 'font_size': 16})
    merge_format3.set_align('vcenter')
    rotate = final_wb.add_format({'rotation': 90, 'align': 'center', 'border': 1, 'font': 'Tahoma', 'font_size': 16})
    rotate.set_align('vcenter')
    ws.set_column('A:A', 6)
    ws.set_column('B:C', 13.5)
    ws.set_column('D:D', 43)
    ws.set_column('E:E', 54)
    ws.set_column('F:F', 9.5)
    ws.set_column('G:G', 18)
    ws.set_column('H:T', 15)
    ws.write('U1', 'Приложение № 2 к Приказу НФ "ПАО "Т Плюс"', format1)
    ws.write('U2', '№___________________________________________ от ____________________________', format1)
    ws.merge_range('A4:U4', 'Техническое задание на поставку ', merge_format2)
    ws.merge_range('A5:C5', 'Таблица 1', merge_format3)
    head = ('№ п/п', '№ лота SAP', 'Код МТР SAP', 'Наименование продукции', 'Технические требования к продукции',
            'Ед. изм.', 'Количество ИТОГО')
    col_head = 0
    for element in head:
        ws.merge_range(5, col_head, 6, col_head, element, merge_format1)
        col_head += 1
    months = ('Декабрь 2020г', 'Январь 2021г', 'Февраль 2021г', 'Март 2021г', 'Апрель 2021г', 'Май 2021г', 'Июнь 2021г',
              'Июль 2021г', 'Август 2021г', 'Сентябрь 2021г', 'Октябрь 2021г', 'Ноябрь 2021г', 'Декабрь 2021г')
    col_month = 7
    for month in months:
        ws.write_string(6, col_month, month, rotate)
        col_month += 1
    ws.merge_range('H6:T6', 'Срок поставки', merge_format1)
    ws.merge_range('U6:U7', 'Грузополучатель', merge_format1)


def consignee(factory_id, lst_len, ws, row_num=7):
    format1 = final_wb.add_format({'align': 'center', 'font': 'Tahoma', 'font_size': 16, 'border': True,
                                   'text_wrap': True})
    ws.set_column('U:U', 46)
    addresses = {'7Q11': 'Сормовская ТЭЦ, 603950, г. Нижний Новгород, ул. Коминтерна, д. 45',
                 '7Q31': 'Новогорьковская ТЭЦ, 6076560, Нижегородская обл, г. Кстово, промзона',
                 '7Q41': 'Дзержинская ТЭЦ , 606000 Нижегородская область, г. Дзержинск, промзона',
                 '7Q91': 'Исполнительный аппарат, 603005, г. Нижний Новгород, Алексеевская 10/16 БЦ "Лобачевский Плаза"',
                 '7Q61': 'Кстовские тепловые сети, Нижегородская обл., г. Кстово, ул. Шохина, 1 корп. 2',
                 '7QB1': 'Дзержинские тепловые сети, Нижегородская обл., г. Дзержинск, ул. Октябрьская, д. 84'}
    i = 0
    row = row_num
    while i <= lst_len:
        ws.write_string(f'U{row}', addresses[f'{factory_id}'], format1)
        i += 1
        row += 1
    return row - 1, row_num


def total(factory_id, ws, row_num, prev_row):
    totals = {'7Q11': 'Итого по Сормовской ТЭЦ', '7Q31': 'Итого по Новогорьковской ТЭЦ',
              '7Q41': 'Итого по Дзержинской ТЭЦ', '7Q91': 'Итого по исполнительному аппарату',
              '7Q61': 'Итого по Кстовским тепловым сетям', '7QB1': 'Итого по Дзержинским тепловые сети'}
    format_total = final_wb.add_format(
        {'bold': True, 'border': 1, 'align': 'center', 'font': 'Tahoma', 'font_size': 16})
    format_total_num = final_wb.add_format({'bold': True, 'border': 1, 'align': 'center', 'num_format': '#,##0.00',
                                            'font': 'Tahoma', 'font_size': 16})
    ws.merge_range(row_num, 0, row_num, 5, totals[f'{factory_id}'], format_total)
    cells = 'GHIJKLMNOPQRST'
    i = 0
    col = 6
    while col <= 19:
        ws.write_formula(row_num, col, f'=SUM({cells[i]}{row_num}:{cells[i]}{prev_row + 1})', format_total_num)
        i += 1
        col += 1
    ws.write(f'U{row_num + 1}', None, format_total)


def make_tail(ws, factory_id, row_num):
    merge_format1 = final_wb.add_format({'align': 'center', 'font': 'Tahoma', 'font_size': 16})
    merge_format1.set_align('vcenter')
    merge_format2 = final_wb.add_format({'align': 'center', 'border': 1, 'font': 'Tahoma', 'font_size': 16,
                                         'text_wrap': True})
    merge_format2.set_align('vcenter')
    merge_format3 = final_wb.add_format({'align': 'left', 'border': 1, 'font': 'Tahoma', 'font_size': 16,
                                         'text_wrap': True})
    merge_format3.set_align('vcenter')
    ws.merge_range(f'A{row_num}:C{row_num}', 'Таблица 2', merge_format1)
    ws.write_string(f'A{row_num + 1}', '№ п/п', merge_format2)
    ws.merge_range(f'B{row_num + 1}:D{row_num + 1}', 'Показатель', merge_format2)
    ws.merge_range(f'E{row_num + 1}:U{row_num + 1}', 'Описание', merge_format2)
    ws.merge_range(f'A{row_num + 2}:A{row_num + 6}', 1, merge_format2)
    ws.merge_range(f'B{row_num + 2}:D{row_num + 6}', 'Условия поставки и отгрузке, требования к упаковке',
                   merge_format3)
    ws.merge_range(f'E{row_num + 2}:U{row_num + 2}', 'Продукция должна быть маркирована и упакована в упаковку, '
                                                     'обеспечивающую сохранность продукции от порчи, повреждений при '
                                                     'транспортировании всеми видами транспорта, перегрузке, хранении, '
                                                     'согласно ГОСТ, ТУ, ОСТ. Тара (упаковка) возврату не подлежит.\n'
                                                     'Особые требования к упаковке: нет.', merge_format3)
    ws.set_row(row_num + 1, 60)
    addresses = {'7Q11': 'Сормовская ТЭЦ, 603950, г. Нижний Новгород, ул. Коминтерна, д. 45',
                 '7Q31': 'Новогорьковская ТЭЦ, 6076560, Нижегородская обл, г. Кстово, промзона',
                 '7Q41': 'Дзержинская ТЭЦ , 606000, Нижегородская область, г. Дзержинск, промзона',
                 '7Q91': 'Исполнительный аппарат, 603005, г. Нижний Новгород, Алексеевская 10/16 БЦ "Лобачевский Плаза"',
                 '7Q61': 'Кстовские тепловые сети, Нижегородская обл., г. Кстово, ул. Шохина, 1 корп. 2',
                 '7QB1': 'Дзержинские тепловые сети, Нижегородская обл., г. Дзержинск, ул. Октябрьская, д. 84'}
    ws.merge_range(f'E{row_num + 3}:U{row_num + 3}', 'Поставка осуществляется путем отгрузок продукции автомобильным '
                                                     'транспортом силами и за счет Поставщика до склада Грузополучателя'
                                                     ' по адресу:\n' f'{addresses[factory_id]}', merge_format3)
    ws.set_row(row_num + 2, 46)
    ws.merge_range(f'E{row_num + 4}:U{row_num + 4}', 'Покупатель вправе отказаться от приемки Товара, поставка которого'
                                                     ' просрочена, в соответствии с условиями договора.', merge_format3)
    ws.merge_range(f'E{row_num + 5}:U{row_num + 5}', 'Одновременно с Продукцией Поставщик передает Покупателю следующие'
                                                     ' сопроводительные документы:\n- Оригинал товарной накладной '
                                                     'унифицированной формы ТОРГ-12/УПД – 2 (два) экземпляра;\n'
                                                     '- Оригинал счета-фактуры на отгруженную Продукцию, оформленного в'
                                                     ' соответствии со статьей 169 НК РФ – 1 (один) экземпляр (оригинал)'
                                                     ' (в случае предоставления УПД счет-фактура не предоставляется);\n'
                                                     '- Оригинал/заверенную копию паспорта изготовителя на поставляемую'
                                                     ' Продукцию;\n- Оригинал/заверенную копию инструкции по '
                                                     'эксплуатации Продукции;\n- Оригинал/заверенную копию документа, '
                                                     'подтверждающего качество поставляемой Продукции (сертификат '
                                                     'качества завода-изготовителя или Поставщика, сертификат '
                                                     'происхождения товара по форме СТ-1, протокол испытаний Продукции '
                                                     'на заводе-изготовителе и т.д.', merge_format3)
    ws.set_row(row_num + 4, 148.20)
    ws.merge_range(f'E{row_num + 6}:U{row_num + 6}', 'Дополнительные требования (наличие шеф-монтажа, послепродажного '
                                                     'технического обслуживания и т. п.): нет', merge_format3)
    ws.merge_range(f'A{row_num + 7}:A{row_num + 10}', 2, merge_format2)
    ws.merge_range(f'B{row_num + 7}:D{row_num + 10}', 'Требования к качеству, гарантийному сроку', merge_format3)
    ws.merge_range(f'E{row_num + 7}:U{row_num + 7}', 'Продукция должна соответствовать обязательным техническим '
                                                     'правилам (ГОСТ, ТУ, РД и др), чертежу, иным техническим '
                                                     'требованиям к продукции, указанным в Таблице 1 Технического '
                                                     'задания.', merge_format3)
    ws.merge_range(f'E{row_num + 8}:U{row_num + 8}', 'Продукция должна быть новым Товаром, который не был в '
                                                     'употреблении, ремонте, в том числе, который не был восстановлен, '
                                                     'у которого не была осуществлена замена составных частей, не были '
                                                     'восстановлены потребительские свойства.', merge_format3)
    ws.set_row(row_num + 7, 40)
    ws.merge_range(f'E{row_num + 9}:U{row_num + 9}', 'В отношении поставляемой продукции Поставщиком устанавливается '
                                                     'гарантийный срок не менее 12 (двенадцати) месяцев с момента '
                                                     'поставки продукции Покупателю.', merge_format3)
    ws.merge_range(f'E{row_num + 10}:U{row_num + 10}', 'Иное: нет', merge_format3)
    ws.write_number(f'A{row_num + 11}', 3, merge_format2)
    ws.merge_range(f'B{row_num + 11}:D{row_num + 11}', 'Подтверждение соответствия продукции предъявляемым требованиям',
                   merge_format3)
    ws.merge_range(f'E{row_num + 11}:U{row_num + 11}', 'На стадии закупки участниками предоставляются:\n- образец/копия'
                                                       ' сертификата соответствия на продукцию (в случае, если '
                                                       'продукция подлежит обязательной сертификации);\n'
                                                       '- санитарно-эпидемиологическое заключение или декларация о '
                                                       'соответствии.', merge_format3)
    ws.set_row(row_num + 10, 66)
    ws.merge_range(f'A{row_num + 12}:A{row_num + 14}', 4, merge_format2)
    ws.merge_range(f'B{row_num + 12}:D{row_num + 14}', 'Требования к безопасности', merge_format3)
    ws.merge_range(f'E{row_num + 12}:U{row_num + 12}', 'Поставщик должен гарантировать безопасность продукции для '
                                                       'жизни, здоровья, имущества Заказчика и окружающей среды при '
                                                       'обычных условиях его использования, хранения, транспортировки '
                                                       'и утилизации.', merge_format3)
    ws.merge_range(f'E{row_num + 13}:U{row_num + 13}', 'Поставляемый Товар должен быть экологически безопасен, '
                                                       'сертифицирован и по безопасности должен соответствовать '
                                                       'требованиям государственных стандартов, техническим условиям и '
                                                       'действующему законодательству РФ.', merge_format3)
    ws.merge_range(f'E{row_num + 14}:U{row_num + 14}', 'Иное: нет', merge_format3)
    ws.merge_range(f'A{row_num + 15}:A{row_num + 18}', 5, merge_format2)
    ws.merge_range(f'B{row_num + 15}:C{row_num + 18}', 'Иные требования', merge_format3)
    ws.write_string(f'D{row_num + 15}', 'Эквивалент', merge_format3)
    ws.write_string(f'D{row_num + 16}', 'Толеранс (+/-), %', merge_format3)
    ws.write_string(f'D{row_num + 17}', 'Срок службы (расчетный ресурс)', merge_format3)
    ws.write_string(f'D{row_num + 18}', 'Другое', merge_format3)
    ws.merge_range(f'E{row_num + 15}:U{row_num + 15}', 'В рамках проведения закупочной процедуры возможна подача '
                                                       'предложений на эквивалентную продукцию. В этом случае участник '
                                                       'должен предоставить документальное подтверждение, что '
                                                       'предлагаемый Товар является полным эквивалентом по техническим'
                                                       ' и функциональным требованиям, характеристикам.', merge_format3)
    ws.set_row(row_num + 14, 42.6)
    ws.merge_range(f'E{row_num + 16}:U{row_num + 16}', 'Нет', merge_format3)
    ws.merge_range(f'E{row_num + 17}:U{row_num + 17}', None, merge_format3)
    ws.merge_range(f'E{row_num + 18}:U{row_num + 18}', 'Нет', merge_format3)
    return row_num + 20


def signatory(ws, factory_id, row_num):
    signatories = {'7Q11': 'Технический директор - главный инженер Сормовской ТЭЦ __________________________________ '
                           '/А.В. Пиголицын/',
                   '7Q31': 'Технический директор - главный инженер Новогорьковской ТЭЦ  _______________________________'
                           '_________ /Р.Г. Валиуллин/',
                   '7Q41': 'Технический директор - главный инженер Дзержинской ТЭЦ _______________________________ '
                           '/Д.А. Чернядьев/',
                   '7Q91': 'Начальник административно-хозяйственного отдела ___________________________ /С.В.Гузняков/',
                   '7Q61': 'Технический директор - главный инженер Кстовских тепловых сетей ___________________________'
                           ' /А.Н. Тихонов/',
                   '7QB1': 'Технический директор-главный инженер Дзержинских тепловых сетей ___________________________'
                           ' /А.В. Семянов/'}
    format1 = final_wb.add_format({'align': 'left', 'bold': True, 'font': 'Tahoma', 'font_size': 16})
    format1.set_align('bottom')
    ws.set_row(row_num - 1, 67.5)
    ws.write_string(f'A{row_num}', f'{signatories[factory_id]}', format1)


if __name__ == "__main__":
    start = time.time()
    budgets = ('РЕМОНТ', 'ЭКСПЛУАТАЦИЯ', 'ИП_ТПИР')  # перечень статей бюджета
    factories = ('7Q11', '7Q31', '7Q41', '7Q91', '7Q61', '7QB1')  # коды грузополучателей
    crs = {'ЦРС ННовг Цех': '7Q11', 'ЦРС Кстово Цехов': '7Q31', 'ЦРС Дзержинск Цехов': '7Q41',
           'НжФ ЦРС ТСКстово Цех': '7Q61', 'НжФ ЦРС ТСДзер Цех': '7QB1'}
    data = pd.read_excel('export.xlsx', sheet_name='Sheet1')
    data['Дата поставки'] = data['Дата поставки'].dt.strftime('%Y/%m')
    data.rename(columns={'Раздел ГКПЗ': 'Раздел_ГКПЗ'}, inplace=True)
    data['Завод'].replace(['7Q71', '7Q81', '7QA1'], '7Q61', inplace=True)
    data['Завод'].replace('7QC1', '7QB1', inplace=True)
    data['Раздел_ГКПЗ'].replace('ИП ТПИР', 'ИП_ТПИР', inplace=True)
    data['Завод'] = data['Наименование МВЗ'].map(crs).fillna(data['Завод'])
    year_month = ('2019/12', '2020/01', '2020/02', '2020/03', '2020/04', '2020/05', '2020/06', '2020/07', '2020/08',
                  '2020/09', '2020/10', '2020/11', '2020/12')
    empty_rows = []
    for element in year_month:
        empty_rows.append({'Раздел_ГКПЗ': '', 'Завод': '', 'Номер лота': '', '№ материала': '',
                           'Краткий текст позиции': '', 'Дата поставки': element, 'ЕИ': '', 'Количество': ''})
    data = data.append(empty_rows, ignore_index=True)
    data_pt = pd.pivot_table(data, index=['Раздел_ГКПЗ', 'Завод', 'Номер лота', '№ материала', 'Краткий текст позиции',
                                          'ЕИ'], values=['Количество'], columns=['Дата поставки'],
                             aggfunc=np.sum).sort_values(by=['Краткий текст позиции'])
    tables = [pivot(factories, budgets[0]), pivot(factories, budgets[1]), pivot(factories, budgets[2])]
    for table in tables:
        for pivots in table:
            pivots.to_excel('temp.xlsx', merge_cells=False)
            temp_wb = load_workbook(filename='temp.xlsx', data_only=True, read_only=True)
            temp_ws = temp_wb.active
            factory_id = temp_ws['B2'].value
            budget_name = temp_ws['A2'].value
            with xl.Workbook(f'ТЗ_{factory_id}_{budget_name}.xlsx') as final_wb:
                final_ws = final_wb.add_worksheet(f'{factory_id}')
                final_ws.set_landscape()
                final_ws.set_paper(9)
                final_ws.fit_to_pages(1, 0)
                final_ws.set_zoom(60)
                make_head(final_ws)
                add_big_table(big_table(temp_ws), final_ws)
                curr_row_num = consignee(factory_id, len(big_table(temp_ws)), final_ws)
                total(factory_id, final_ws, curr_row_num[0], curr_row_num[1])
                row_for_sign = make_tail(final_ws, factory_id, curr_row_num[0] + 3)
                signatory(final_ws, factory_id, row_for_sign)
            temp_wb.close()
    os.remove('temp.xlsx')
    print('Lead time: {:.2f} secs.'.format(time.time() - start))
