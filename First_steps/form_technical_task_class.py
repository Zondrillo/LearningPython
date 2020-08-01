import pandas as pd
import numpy as np
import time
import os
import tempfile as tf
import xlsxwriter as xl
from openpyxl import load_workbook
from datetime import datetime


def pivot(file_name):  # создаёт списки сводных таблиц для каждого грузополучателя, в соответствии с бюджетом
    factories = ('7Q11', '7Q31', '7Q41', '7Q61', '7Q91', '7QB1')  # коды грузополучателей
    budgets = ('РЕМОНТ', 'ЭКСПЛУАТАЦИЯ', 'ИП_ТПИР')  # перечень статей бюджета
    crs = {'ЦРС ННовг Цех': '7Q11', 'ЦРС Кстово Цехов': '7Q31', 'ЦРС Дзержинск Цехов': '7Q41',  # Наименования МВЗ
           'НжФ ЦРС ТСКстово Цех': '7Q61', 'НжФ ЦРС ТСДзер Цех': '7QB1'}
    curr_year = datetime.now().year
    next_year = curr_year + 1
    year_month = (f'{curr_year}/12', f'{next_year}/01', f'{next_year}/02', f'{next_year}/03', f'{next_year}/04',
                  f'{next_year}05', f'{next_year}/06', f'{next_year}/07', f'{next_year}/08', f'{next_year}/09',
                  f'{next_year}/10', f'{next_year}/11', f'{next_year}/12')  # годы/месяцы поставки
    data = pd.read_excel(file_name, sheet_name='Sheet1')
    data['Дата поставки'] = data['Дата поставки'].dt.strftime('%Y/%m')  # преобразование дат в формат ГГГГ/ММ
    data.rename(columns={'Раздел ГКПЗ': 'Раздел_ГКПЗ'}, inplace=True)
    data['Завод'].replace(['7Q71', '7Q81', '7QA1'], '7Q61', inplace=True)
    data['Завод'].replace('7QC1', '7QB1', inplace=True)
    data['Раздел_ГКПЗ'].replace('ИП ТПИР', 'ИП_ТПИР', inplace=True)
    data['Завод'] = data['Наименование МВЗ'].map(crs).fillna(data['Завод'])  # распределение позиций ЦРС по заводам
    empty_rows = []
    for element in year_month:
        empty_rows.append({'Раздел_ГКПЗ': '', 'Завод': '', 'Номер лота': '', '№ материала': '',
                           'Краткий текст позиции': '', 'Дата поставки': element, 'ЕИ': '', 'Количество': ''})
    data = data.append(empty_rows, ignore_index=True)  # фиксируем диапазон дат поставки
    data_pt = pd.pivot_table(data, index=['Раздел_ГКПЗ', 'Завод', 'Номер лота', '№ материала', 'Краткий текст позиции',
                                          'ЕИ'], values=['Количество'], columns=['Дата поставки'],
                             aggfunc=np.sum).sort_values(by=['Краткий текст позиции'])  # формируем общую сводную таблицу
    df = []
    for budget in budgets:  # создаём отдельные сводные таблицы для каждого завода и раздела ГКПЗ
        for factory in factories:
            temp_pivot = data_pt.query(f'Завод == ["{factory}"] & Раздел_ГКПЗ == ["{budget}"]')
            if temp_pivot.size != 0:
                df.append(temp_pivot)
    return df


class FormTechTaskSep:
    addresses = {'7Q11': 'Сормовская ТЭЦ, 603950, г. Нижний Новгород, ул. Коминтерна, д. 45',  # адреса грузополучателей
                 '7Q31': 'Новогорьковская ТЭЦ, 6076560, Нижегородская обл, г. Кстово, промзона',
                 '7Q41': 'Дзержинская ТЭЦ , 606000 Нижегородская область, г. Дзержинск, промзона',
                 '7Q91': 'Исполнительный аппарат, 603005, г. Нижний Новгород, Алексеевская 10/16 БЦ "Лобачевский Плаза"',
                 '7Q61': 'Кстовские тепловые сети, Нижегородская обл., г. Кстово, ул. Шохина, 1 корп. 2',
                 '7QB1': 'Дзержинские тепловые сети, Нижегородская обл., г. Дзержинск, ул. Октябрьская, д. 84'}

    def __init__(self, some_table):
        self.file_path = tf.mktemp(suffix='.xlsx', dir='')  # создаём временный файл для записи сводной таблицы одного завода
        some_table.to_excel(self.file_path, merge_cells=False)  # преобразовываем сводную таблицу в формат excel
        self.temp_wb = load_workbook(filename=self.file_path, data_only=True)  # получаем данные из excel-файла со сводной таблицей
        self.temp_ws = self.temp_wb.active  # выбираем единственный временный лист в excel-файле
        self.factory_id = self.temp_ws['B2'].value  # запоминаем id завода, с которым работаем в данный момент
        self.budget_name = self.temp_ws['A2'].value  # запоминаем раздел ГКПЗ с которым работаем в данный момент
        self.final_wb = xl.Workbook(f'ТЗ_{self.factory_id}_{self.budget_name}.xlsx')  # создаём конечный excel-файл, в который будем записывать данные
        self.final_ws = self.final_wb.add_worksheet(f'{self.factory_id}')  # добавляем лист, в который будем записывать данные
        self.final_ws.set_landscape()  # альбомная ориентация
        self.final_ws.set_paper(9)  # формат А4
        self.final_ws.fit_to_pages(1, 0)  # вписать все столбцы на одну страницу
        self.final_ws.set_zoom(60)  # установить масштаб 60%

    def big_table(self):  # получаем все необходимые данные из сводной таблицы
        lst = []
        i = 1
        for row in self.temp_ws.iter_rows(min_row=2, max_row=self.temp_ws.max_row, min_col=3, max_col=19):
            tmp_lst = [i]
            j = 0
            while j < len(row):
                if j == 3:
                    tmp_lst.append(None)
                tmp_lst.append(row[j].value)
                j += 1
            lst.append(tmp_lst)
            i += 1
        return lst

    def make_middle(self, lst, factory_id, row_num=8):  # добавляем данные в таблицу 1 ТЗ
        r_num = row_num
        format1 = self.final_wb.add_format({'align': 'center', 'border': True, 'font': 'Tahoma', 'font_size': 16,
                                            'text_wrap': True})
        format1.set_align('vcenter')
        quantity_format = self.final_wb.add_format({'num_format': '#,##0.00', 'align': 'center', 'border': True,
                                                    'font': 'Tahoma', 'font_size': 16})
        quantity_format.set_align('vcenter')
        format_total = self.final_wb.add_format({'bold': True, 'border': 1, 'align': 'center', 'font': 'Tahoma',
                                                 'font_size': 16})
        format_total_num = self.final_wb.add_format({'bold': True, 'border': 1, 'align': 'center', 'num_format':
                                                     '#,##0.00', 'font': 'Tahoma', 'font_size': 16})
        self.final_ws.set_column('U:U', 46)
        for row in lst:
            self.final_ws.write_row(f'A{r_num}', row[:6], format1)
            self.final_ws.write_formula(r_num - 1, 6, f'=SUM(H{r_num}:T{r_num})', quantity_format)
            self.final_ws.write_row(f'H{r_num}', row[6:], quantity_format)
            self.final_ws.write_string(f'U{r_num}', self.addresses[f'{factory_id}'], format1)
            r_num += 1
        totals = {'7Q11': 'Итого по Сормовской ТЭЦ', '7Q31': 'Итого по Новогорьковской ТЭЦ',
                  '7Q41': 'Итого по Дзержинской ТЭЦ', '7Q91': 'Итого по исполнительному аппарату',
                  '7Q61': 'Итого по Кстовским тепловым сетям', '7QB1': 'Итого по Дзержинским тепловые сети'}
        self.final_ws.merge_range(r_num - 1, 0, r_num - 1, 5, totals[f'{factory_id}'], format_total)
        cells = 'GHIJKLMNOPQRST'
        i = 0
        col = 6
        while col <= 19:
            self.final_ws.write_formula(r_num - 1, col, f'=SUM({cells[i]}{row_num}:{cells[i]}{r_num - 1})',
                                        format_total_num)
            i += 1
            col += 1
        self.final_ws.write(f'U{r_num}', None, format_total)
        return r_num

    def make_head(self):  # делаем шапку ТЗ
        format1 = self.final_wb.add_format({'align': 'right', 'italic': True, 'font': 'Tahoma', 'font_size': 16})
        merge_format1 = self.final_wb.add_format({'align': 'center', 'border': 1, 'font': 'Tahoma', 'font_size': 16,
                                                  'text_wrap': True})
        merge_format1.set_align('vcenter')
        merge_format2 = self.final_wb.add_format({'align': 'center', 'bold': True, 'font': 'Tahoma', 'font_size': 16})
        merge_format2.set_align('vcenter')
        merge_format3 = self.final_wb.add_format({'align': 'center', 'font': 'Tahoma', 'font_size': 16})
        merge_format3.set_align('vcenter')
        rotate = self.final_wb.add_format({'rotation': 90, 'align': 'center', 'border': 1, 'font': 'Tahoma',
                                           'font_size': 16})
        rotate.set_align('vcenter')
        self.final_ws.set_column('A:A', 6)
        self.final_ws.set_column('B:C', 13.5)
        self.final_ws.set_column('D:D', 43)
        self.final_ws.set_column('E:E', 54)
        self.final_ws.set_column('F:F', 9.5)
        self.final_ws.set_column('G:G', 18)
        self.final_ws.set_column('H:T', 15)
        self.final_ws.write('U1', 'Приложение № 2 к Приказу НФ "ПАО "Т Плюс"', format1)
        self.final_ws.write('U2', '№___________________________________________ от ____________________________',
                            format1)
        self.final_ws.merge_range('A4:U4', 'Техническое задание на поставку ', merge_format2)
        self.final_ws.merge_range('A5:C5', 'Таблица 1', merge_format3)
        head = ('№ п/п', '№ лота SAP', 'Код МТР SAP', 'Наименование продукции', 'Технические требования к продукции',
                'Ед. изм.', 'Количество ИТОГО')
        col_head = 0
        for element in head:
            self.final_ws.merge_range(5, col_head, 6, col_head, element, merge_format1)
            col_head += 1
        months = ('Декабрь 2020г', 'Январь 2021г', 'Февраль 2021г', 'Март 2021г', 'Апрель 2021г', 'Май 2021г',
                  'Июнь 2021г', 'Июль 2021г', 'Август 2021г', 'Сентябрь 2021г', 'Октябрь 2021г', 'Ноябрь 2021г',
                  'Декабрь 2021г')
        col_month = 7
        for month in months:
            self.final_ws.write_string(6, col_month, month, rotate)
            col_month += 1
        self.final_ws.merge_range('H6:T6', 'Срок поставки', merge_format1)
        self.final_ws.merge_range('U6:U7', 'Грузополучатель', merge_format1)

    def make_tail(self, factory_id, row_num):  # Вставляем таблицу 2 ТЗ
        merge_format1 = self.final_wb.add_format({'align': 'center', 'font': 'Tahoma', 'font_size': 16})
        merge_format1.set_align('vcenter')
        merge_format2 = self.final_wb.add_format({'align': 'center', 'border': 1, 'font': 'Tahoma', 'font_size': 16,
                                                  'text_wrap': True})
        merge_format2.set_align('vcenter')
        merge_format3 = self.final_wb.add_format({'align': 'left', 'border': 1, 'font': 'Tahoma', 'font_size': 16,
                                                  'text_wrap': True})
        merge_format3.set_align('vcenter')
        self.final_ws.merge_range(f'A{row_num}:C{row_num}', 'Таблица 2', merge_format1)
        self.final_ws.write_string(f'A{row_num + 1}', '№ п/п', merge_format2)
        self.final_ws.merge_range(f'B{row_num + 1}:D{row_num + 1}', 'Показатель', merge_format2)
        self.final_ws.merge_range(f'E{row_num + 1}:U{row_num + 1}', 'Описание', merge_format2)
        self.final_ws.merge_range(f'A{row_num + 2}:A{row_num + 6}', 1, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 2}:D{row_num + 6}',
                                  'Условия поставки и отгрузке, требования к упаковке', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 2}:U{row_num + 2}',
                                  'Продукция должна быть маркирована и упакована в упаковку, обеспечивающую сохранность'
                                  ' продукции от порчи, повреждений при транспортировании всеми видами транспорта, '
                                  'перегрузке, хранении, согласно ГОСТ, ТУ, ОСТ. Тара (упаковка) возврату не подлежит.'
                                  '\nОсобые требования к упаковке: нет.', merge_format3)
        self.final_ws.set_row(row_num + 1, 60)
        self.final_ws.merge_range(f'E{row_num + 3}:U{row_num + 3}',
                                  'Поставка осуществляется путем отгрузок продукции автомобильным транспортом силами и '
                                  'за счет Поставщика до склада Грузополучателя по адресу:\n'
                                  f'{self.addresses[factory_id]}', merge_format3)
        self.final_ws.set_row(row_num + 2, 46)
        self.final_ws.merge_range(f'E{row_num + 4}:U{row_num + 4}',
                                  'Покупатель вправе отказаться от приемки Товара, поставка которого просрочена, в '
                                  'соответствии с условиями договора.', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 5}:U{row_num + 5}',
                                  'Одновременно с Продукцией Поставщик передает Покупателю следующие cопроводительные '
                                  'документы:\n- Оригинал товарной накладной унифицированной формы ТОРГ-12/УПД – 2 '
                                  '(два) экземпляра;\n- Оригинал счета-фактуры на отгруженную Продукцию, оформленного '
                                  'в соответствии со статьей 169 НК РФ – 1 (один) экземпляр (оригинал) (в случае '
                                  'предоставления УПД счет-фактура не предоставляется);\n- Оригинал/заверенную копию '
                                  'паспорта изготовителя на поставляемую Продукцию;\n- Оригинал/заверенную копию '
                                  'инструкции по эксплуатации Продукции;\n-Оригинал/заверенную копию документа, '
                                  'подтверждающего качество поставляемой Продукции (сертификат качества '
                                  'завода-изготовителя или Поставщика, сертификат происхождения товара по форме СТ-1, '
                                  'протокол испытаний Продукции на заводе-изготовителе и т.д.', merge_format3)
        self.final_ws.set_row(row_num + 4, 148.20)
        self.final_ws.merge_range(f'E{row_num + 6}:U{row_num + 6}',
                                  'Дополнительные требования (наличие шеф-монтажа, послепродажного технического '
                                  'обслуживания и т. п.): нет', merge_format3)
        self.final_ws.merge_range(f'A{row_num + 7}:A{row_num + 10}', 2, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 7}:D{row_num + 10}', 'Требования к качеству, гарантийному сроку',
                                  merge_format3)
        self.final_ws.merge_range(f'E{row_num + 7}:U{row_num + 7}',
                                  'Продукция должна соответствовать обязательным техническим правилам (ГОСТ, ТУ, РД и '
                                  'др), чертежу, иным техническим требованиям к продукции, указанным в Таблице 1 '
                                  'Технического задания.', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 8}:U{row_num + 8}',
                                  'Продукция должна быть новым Товаром, который не был в употреблении, ремонте, в том '
                                  'числе, который не был восстановлен, у которого не была осуществлена замена составных'
                                  ' частей, не были восстановлены потребительские свойства.', merge_format3)
        self.final_ws.set_row(row_num + 7, 40)
        self.final_ws.merge_range(f'E{row_num + 9}:U{row_num + 9}',
                                  'В отношении поставляемой продукции Поставщиком устанавливается гарантийный срок не '
                                  'менее 12 (двенадцати) месяцев с момента поставки продукции Покупателю.',
                                  merge_format3)
        self.final_ws.merge_range(f'E{row_num + 10}:U{row_num + 10}', 'Иное: нет', merge_format3)
        self.final_ws.write_number(f'A{row_num + 11}', 3, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 11}:D{row_num + 11}', 'Подтверждение соответствия продукции '
                                                                      'предъявляемым требованиям', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 11}:U{row_num + 11}',
                                  'На стадии закупки участниками предоставляются:\n- образец/копия сертификата '
                                  'соответствия на продукцию (в случае, если продукция подлежит обязательной '
                                  'сертификации);\n- санитарно-эпидемиологическое заключение или декларация о '
                                  'соответствии.', merge_format3)
        self.final_ws.set_row(row_num + 10, 66)
        self.final_ws.merge_range(f'A{row_num + 12}:A{row_num + 14}', 4, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 12}:D{row_num + 14}', 'Требования к безопасности', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 12}:U{row_num + 12}',
                                  'Поставщик должен гарантировать безопасность продукции для жизни, здоровья, имущества'
                                  ' Заказчика и окружающей среды при обычных условиях его использования, хранения, '
                                  'транспортировки и утилизации.', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 13}:U{row_num + 13}',
                                  'Поставляемый Товар должен быть экологически безопасен, сертифицирован и по '
                                  'безопасности должен соответствовать требованиям государственных стандартов, '
                                  'техническим условиям и действующему законодательству РФ.', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 14}:U{row_num + 14}', 'Иное: нет', merge_format3)
        self.final_ws.merge_range(f'A{row_num + 15}:A{row_num + 18}', 5, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 15}:C{row_num + 18}', 'Иные требования', merge_format3)
        self.final_ws.write_string(f'D{row_num + 15}', 'Эквивалент', merge_format3)
        self.final_ws.write_string(f'D{row_num + 16}', 'Толеранс (+/-), %', merge_format3)
        self.final_ws.write_string(f'D{row_num + 17}', 'Срок службы (расчетный ресурс)', merge_format3)
        self.final_ws.write_string(f'D{row_num + 18}', 'Другое', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 15}:U{row_num + 15}',
                                  'В рамках проведения закупочной процедуры возможна подача предложений на '
                                  'эквивалентную продукцию. В этом случае участник должен предоставить документальное '
                                  'подтверждение, что предлагаемый Товар является полным эквивалентом по техническим'
                                  ' и функциональным требованиям, характеристикам.', merge_format3)
        self.final_ws.set_row(row_num + 14, 42.6)
        self.final_ws.merge_range(f'E{row_num + 16}:U{row_num + 16}', 'Нет', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 17}:U{row_num + 17}', None, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 18}:U{row_num + 18}', 'Нет', merge_format3)
        return row_num + 20

    def signatory(self, factory_id, row_num):  # добавляем подписанта
        signatories = {'7Q11': 'Технический директор - главный инженер Сормовской ТЭЦ ________________________________'
                               '__ /А.В. Пиголицын/',
                       '7Q31': 'Технический директор - главный инженер Новогорьковской ТЭЦ  ___________________________'
                               '_____________ /Р.Г. Валиуллин/',
                       '7Q41': 'Технический директор - главный инженер Дзержинской ТЭЦ _______________________________ '
                               '/Д.А. Чернядьев/',
                       '7Q91': 'Начальник административно-хозяйственного отдела ___________________________ '
                               '/С.В.Гузняков/',
                       '7Q61': 'Технический директор - главный инженер Кстовских тепловых сетей _______________________'
                               '____ /А.Н. Тихонов/',
                       '7QB1': 'Технический директор-главный инженер Дзержинских тепловых сетей _______________________'
                               '____ /А.В. Семянов/'}
        format1 = self.final_wb.add_format({'align': 'left', 'bold': True, 'font': 'Tahoma', 'font_size': 16})
        format1.set_align('bottom')
        self.final_ws.set_row(row_num - 1, 67.5)
        self.final_ws.write_string(f'A{row_num}', f'{signatories[factory_id]}', format1)

    def form(self):  # формируем ТЗ
        self.make_head()
        curr_row_num = self.make_middle(self.big_table(), self.factory_id)
        row_for_sign = self.make_tail(self.factory_id, curr_row_num + 2)
        self.signatory(self.factory_id, row_for_sign)
        self.temp_wb.close()
        self.final_wb.close()
        os.remove(self.file_path)


if __name__ == "__main__":
    start = time.time()
    tables = pivot('export.xlsx')
    for table in tables:
        FormTechTaskSep(table).form()
    print('Lead time: {:.2f} secs.'.format(time.time() - start))
